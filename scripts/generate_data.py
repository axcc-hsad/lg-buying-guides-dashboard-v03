#!/usr/bin/env python3

import argparse
import json
import math
import re
import sys
from collections import defaultdict
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / ".vendor"))

from openpyxl import load_workbook  # type: ignore

from workbook_config import (
    COUNTRY_CANONICAL_MAP,
    COUNTRY_COUNT_GROUP_MAP,
    IGNORED_SHEETS,
    METRIC_ALIASES,
    MONTH_NAMES,
    NARRATIVE_TEMPLATES,
    SHEET_LAYOUT,
)


def canonical_country(raw):
    if raw is None:
        return raw
    normalized = str(raw).strip().replace("_", "-").upper()
    return COUNTRY_CANONICAL_MAP.get(normalized, normalized)


def country_count_group(raw):
    canonical = canonical_country(raw)
    return COUNTRY_COUNT_GROUP_MAP.get(canonical, canonical)


def slugify(value):
    text = str(value or "").strip().lower()
    text = text.replace("&", " and ")
    text = text.replace("%", " percent ")
    text = text.replace("/", "_")
    text = text.replace("-", "_")
    text = re.sub(r"\([^)]*\)", lambda m: "_" + m.group(0)[1:-1] + "_", text)
    text = re.sub(r"[^a-z0-9가-힣_]+", "_", text)
    text = re.sub(r"_+", "_", text).strip("_")
    return text


def leading_indent(label):
    return len(str(label or "")) - len(str(label or "").lstrip())


def is_number(value):
    return isinstance(value, (int, float)) and not isinstance(value, bool) and not math.isnan(value)


def to_number(value):
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        if isinstance(value, float) and math.isnan(value):
            return None
        return value
    if isinstance(value, str):
        cleaned = value.strip().replace(",", "")
        if not cleaned:
            return None
        try:
            return float(cleaned)
        except ValueError:
            return None
    return None


def detect_month_columns(ws):
    for row in range(1, min(ws.max_row, SHEET_LAYOUT["month_scan_rows"]) + 1):
        matches = {}
        for col in range(1, ws.max_column + 1):
            value = ws.cell(row=row, column=col).value
            if value in MONTH_NAMES:
                matches[col] = value
        if len(matches) >= 2:
            return row, matches
    raise ValueError(f"Could not detect month columns in sheet {ws.title}")


def detect_year_blocks(ws):
    blocks = []
    max_row = min(ws.max_row, SHEET_LAYOUT["month_scan_rows"])
    for row in range(1, max_row + 1):
        year_value = ws.cell(row=row, column=2).value
        if not isinstance(year_value, (int, float)):
            continue
        year = int(year_value)
        if year < 2000 or year > 2100:
            continue
        matches = {}
        for col in range(1, ws.max_column + 1):
            value = ws.cell(row=row, column=col).value
            if value in MONTH_NAMES:
                matches[col] = value
        if len(matches) >= 2:
            blocks.append((year, row, matches))
    return blocks


def month_name_from_korean(text):
    match = re.search(r"(\d{1,2})월", str(text or ""))
    if not match:
        return None
    month_no = int(match.group(1))
    if 1 <= month_no <= 12:
        return MONTH_NAMES[month_no - 1]
    return None


def months_with_data(records, month_columns):
    available = []
    for col, month in month_columns.items():
        if any(to_number(ws_row[col]) is not None for ws_row in records):
            available.append(month)
    return available


def parse_sheet_name(sheet_name):
    parts = sheet_name.split("_", 1)
    if len(parts) != 2:
        raise ValueError(f"Unsupported sheet name format: {sheet_name}")
    category, country = parts
    return category, canonical_country(country)


def alias_key(raw_key):
    for alias, patterns in METRIC_ALIASES.items():
        if any(pattern in raw_key for pattern in patterns):
            return alias
    return None


def parse_sheet_insights(ws):
    insights = []
    capture = False
    for row in range(1, min(ws.max_row, 80) + 1):
        left = ws.cell(row=row, column=2).value
        right = ws.cell(row=row, column=3).value
        left_text = str(left or "").strip()
        right_text = str(right or "").strip()
        if "<Data Insights>" in left_text:
            capture = True
            continue
        if not capture:
            continue
        if not left_text and not right_text:
            if insights:
                break
            continue
        if left_text.startswith(("1)", "2)", "3)")):
            insight = left_text + (f" {right_text}" if right_text else "")
            insights.append(insight.strip())
    return insights


def extract_source_rows(values_ws, formula_ws):
    values = {}
    max_row = max(values_ws.max_row, formula_ws.max_row)
    max_col = max(values_ws.max_column, formula_ws.max_column)
    for row in range(1, max_row + 1):
        row_values = {}
        for col in range(1, max_col + 1):
            formula_value = formula_ws.cell(row=row, column=col).value
            data_value = values_ws.cell(row=row, column=col).value
            # Ignore stale cached values when the workbook has no actual formula/value in the raw sheet.
            row_values[col] = data_value if formula_value is not None else None
        values[row] = row_values
    return values


def period_key(year, month):
    return f"{year}-{month}"


def parse_sheet(values_ws, formula_ws, content_type):
    category, country = parse_sheet_name(values_ws.title)
    year_blocks = detect_year_blocks(formula_ws)
    if not year_blocks:
        month_row, month_columns = detect_month_columns(formula_ws)
        inferred_year = 2026
        year_blocks = [(inferred_year, month_row, month_columns)]
    all_rows = extract_source_rows(values_ws, formula_ws)
    merged_metrics = {}
    metrics_by_year = {}
    periods_available = []
    max_row = max(values_ws.max_row, formula_ws.max_row)

    for index, (year, month_row, month_columns) in enumerate(year_blocks):
        metrics = metrics_by_year.setdefault(year, {})
        current_section = None
        empty_streak = 0
        next_year_row = year_blocks[index + 1][1] if index + 1 < len(year_blocks) else max_row + 1

        for row in range(month_row + 1, next_year_row):
            row_month_hits = {
                col: all_rows[row].get(col)
                for col in range(1, max(values_ws.max_column, formula_ws.max_column) + 1)
                if all_rows[row].get(col) in MONTH_NAMES
            }
            if len(row_month_hits) >= 2 and row != month_row:
                break

            section_value = all_rows[row].get(SHEET_LAYOUT["section_column"])
            label_value = all_rows[row].get(SHEET_LAYOUT["label_column"])
            if section_value not in (None, ""):
                current_section = str(section_value).strip()

            month_values = {}
            merged_month_values = {}
            for col, month in month_columns.items():
                numeric = to_number(all_rows[row].get(col))
                if numeric is not None:
                    month_values[month] = numeric
                    merged_month_values[period_key(year, month)] = numeric

            if not label_value and not month_values:
                empty_streak += 1
                if empty_streak >= SHEET_LAYOUT["empty_row_break"]:
                    break
                continue

            empty_streak = 0
            if not label_value or not month_values:
                continue

            raw_key = slugify(label_value)
            if not raw_key:
                continue

            year_record = metrics.get(raw_key)
            if not year_record:
                year_record = {
                    "section": current_section,
                    "label": str(label_value).strip(),
                    "indent": leading_indent(label_value),
                    "monthly": {},
                    "source_sheet": values_ws.title,
                    "source_cell": f"C{row}",
                }
                metrics[raw_key] = year_record
            year_record["monthly"].update(month_values)

            merged_record = merged_metrics.get(raw_key)
            if not merged_record:
                merged_record = {
                    "section": current_section,
                    "label": str(label_value).strip(),
                    "indent": leading_indent(label_value),
                    "monthly": {},
                    "source_sheet": values_ws.title,
                    "source_cell": f"C{row}",
                }
                merged_metrics[raw_key] = merged_record
            merged_record["monthly"].update(merged_month_values)

            canonical = alias_key(raw_key)
            if canonical:
                if canonical not in metrics:
                    metrics[canonical] = year_record
                if canonical not in merged_metrics:
                    merged_metrics[canonical] = merged_record

        for month in MONTH_NAMES:
            if any(period_key(year, month) in metric["monthly"] for metric in merged_metrics.values()):
                periods_available.append(period_key(year, month))

    return {
        "category": category,
        "country": country,
        "content_type": content_type,
        "metrics_2026": merged_metrics,
        "metrics_2025": metrics_by_year.get(2025, {}),
        "metrics_by_year": {str(year): metrics for year, metrics in metrics_by_year.items()},
        "insights": parse_sheet_insights(values_ws),
        "_periods": periods_available,
    }


def parse_ai_sheet(values_ws, formula_ws, content_type):
    records = []
    current_guide_type = None
    current_month = None
    current_year = None
    country_columns = {}
    max_row = max(values_ws.max_row, formula_ws.max_row)
    max_col = max(values_ws.max_column, formula_ws.max_column)

    for row in range(1, max_row + 1):
        left = formula_ws.cell(row=row, column=2).value
        metric_label = formula_ws.cell(row=row, column=3).value
        left_text = str(left or "").strip()

        if "월 기준" in left_text:
            current_month = month_name_from_korean(left_text)
            current_guide_type = "TV" if left_text.upper().startswith("TV") else "Monitor" if left_text.lower().startswith("monitor") else None
            country_columns = {}
            continue

        if current_month and isinstance(left, (int, float)):
            current_year = int(left)
            country_columns = {}
            for col in range(9, max_col + 1):
                country = formula_ws.cell(row=row, column=col).value
                if not country:
                    continue
                country_columns[col] = canonical_country(country)
            continue

        if not current_month or not current_guide_type or not country_columns or not metric_label:
            continue

        metric_key = alias_key(slugify(metric_label)) or slugify(metric_label)
        if not metric_key:
            continue

        for col, country in country_columns.items():
            formula_value = formula_ws.cell(row=row, column=col).value
            if formula_value is None:
                continue
            value = to_number(values_ws.cell(row=row, column=col).value)
            if value is None:
                continue
            records.append({
                "month": current_month,
                "year": current_year,
                "country": country,
                "guide_type": current_guide_type,
                "content_type": content_type,
                "metric": metric_key,
                "label": str(metric_label).strip(),
                "value": value,
                "source_sheet": values_ws.title,
                "source_cell": f"{chr(64 + col) if col <= 26 else col}{row}",
            })

    months = [record["month"] for record in records if record.get("month") in MONTH_NAMES]
    latest = MONTH_NAMES[max(MONTH_NAMES.index(month) for month in months)] if months else None
    return {
        "latest_month": latest,
        "records": records,
    }


def get_metric_value(metrics, month, *keywords):
    lowered = [keyword.lower() for keyword in keywords]
    for key, metric in metrics.items():
        if all(keyword in key.lower() for keyword in lowered):
            return metric["monthly"].get(month)
    return None


def average(values):
    valid = [value for value in values if value is not None]
    return sum(valid) / len(valid) if valid else None


def pct_change(current, previous):
    if current is None or previous in (None, 0):
        return None
    return ((current - previous) / abs(previous)) * 100


def summarize_sheet(item, latest, prev):
    metrics = item["metrics_2026"]
    sessions = (
        get_metric_value(metrics, latest, "lineup", "session")
        or get_metric_value(metrics, latest, item["category"].lower(), "lineup", "session")
        or get_metric_value(metrics, latest, "session")
    )
    sessions_prev = (
        get_metric_value(metrics, prev, "lineup", "session")
        or get_metric_value(metrics, prev, item["category"].lower(), "lineup", "session")
        or get_metric_value(metrics, prev, "session")
    ) if prev else None
    return {
        "content_type": item["content_type"],
        "sessions": sessions,
        "sessions_prev": sessions_prev,
        "page_views": get_metric_value(metrics, latest, "page_view") or get_metric_value(metrics, latest, item["category"].lower(), "line", "up"),
        "engaged": get_metric_value(metrics, latest, "engaged"),
        "event_clicks": get_metric_value(metrics, latest, "event_click"),
        "event_clicks_prev": get_metric_value(metrics, prev, "event_click") if prev else None,
        "duration": get_metric_value(metrics, latest, "avg_session_duration"),
        "duration_prev": get_metric_value(metrics, prev, "avg_session_duration") if prev else None,
        "plp_conv": get_metric_value(metrics, latest, "plp_conversion"),
        "plp_conv_prev": get_metric_value(metrics, prev, "plp_conversion") if prev else None,
        "product_conv": get_metric_value(metrics, latest, "product_conversion"),
        "purchase_conv": get_metric_value(metrics, latest, "purchase_conversion"),
        "purchase_conv_prev": get_metric_value(metrics, prev, "purchase_conversion") if prev else None,
        "purchase_count": get_metric_value(metrics, latest, "purchase"),
        "organic": get_metric_value(metrics, latest, "organic"),
        "external": get_metric_value(metrics, latest, "external_entrance"),
        "internal": get_metric_value(metrics, latest, "internal"),
        "engagement_rate": get_metric_value(metrics, latest, "engagement_rate"),
        "exit_rate": get_metric_value(metrics, latest, "exit_rate"),
    }


def build_country_reports(parsed_sheets, sheet_summaries, latest, prev):
    reports = {}
    for key, item in parsed_sheets.items():
        country = item["country"]
        summary = sheet_summaries[key]
        current = reports.setdefault(country, {
            "total_sessions": 0,
            "session_change_pct": None,
            "total_clicks": 0,
            "avg_plp_conv": None,
            "avg_purchase_conv": None,
            "total_purchases": 0,
            "avg_duration": None,
            "status": "stable",
            "links": [],
            "analyst_notes": [],
            "per_product": {},
        })
        current["total_sessions"] += summary["sessions"] or 0
        current["total_clicks"] += summary["event_clicks"] or 0
        current["total_purchases"] += summary["purchase_count"] or 0
        current["avg_plp_conv"] = average([current["avg_plp_conv"], summary["plp_conv"]])
        current["avg_purchase_conv"] = average([current["avg_purchase_conv"], summary["purchase_conv"]])
        current["avg_duration"] = average([current["avg_duration"], summary["duration"]])
        current["per_product"][item["category"]] = {
            "sessions": summary["sessions"],
            "page_views": summary["page_views"],
            "plp_conv": summary["plp_conv"],
            "purchase_conv": summary["purchase_conv"],
            "duration": summary["duration"],
            "clicks": summary["event_clicks"],
            "engagement_rate": summary["engagement_rate"],
        }
        if item["insights"]:
            current["analyst_notes"].append({
                "page": item["category"],
                "notes": item["insights"],
            })

    for country, report in reports.items():
        previous_sessions = sum(
            sheet_summaries[key]["sessions_prev"] or 0
            for key, item in parsed_sheets.items()
            if item["country"] == country
        )
        report["session_change_pct"] = pct_change(report["total_sessions"], previous_sessions)
        plp_change = None
        if latest and prev:
            current_plp = report["avg_plp_conv"]
            previous_plp = average([
                sheet_summaries[key]["plp_conv_prev"]
                for key, item in parsed_sheets.items()
                if item["country"] == country
            ])
            plp_change = pct_change(current_plp, previous_plp)
        session_change = report["session_change_pct"] or 0
        plp_change = plp_change or 0
        if session_change >= 5 and plp_change >= 0:
            report["status"] = "growing"
        elif session_change <= -10 or plp_change <= -10:
            report["status"] = "declining"
        else:
            report["status"] = "stable"
    return reports


def build_category_reports(parsed_sheets, sheet_summaries):
    categories = {}
    for key, item in parsed_sheets.items():
        category = item["category"]
        summary = sheet_summaries[key]
        entry = categories.setdefault(category, {
            "total_sessions": 0,
            "avg_plp_conv": None,
            "countries_count": 0,
            "ranked_sessions": [],
            "ranked_plp": [],
            "ranked_duration": [],
        })
        entry["total_sessions"] += summary["sessions"] or 0
        entry["avg_plp_conv"] = average([entry["avg_plp_conv"], summary["plp_conv"]])
        entry["ranked_sessions"].append([item["country"], summary["sessions"] or 0])
        entry["ranked_plp"].append([item["country"], summary["plp_conv"] or 0])
        entry["ranked_duration"].append([item["country"], summary["duration"] or 0])

    for entry in categories.values():
        entry["countries_count"] = len({country_count_group(country) for country, _ in entry["ranked_sessions"]})
        entry["ranked_sessions"].sort(key=lambda row: row[1], reverse=True)
        entry["ranked_plp"].sort(key=lambda row: row[1], reverse=True)
        entry["ranked_duration"].sort(key=lambda row: row[1], reverse=True)
    return categories


def build_executive(parsed_sheets, sheet_summaries, latest, prev):
    summaries = list(sheet_summaries.values())
    total_sessions = sum(item["sessions"] or 0 for item in summaries)
    total_sessions_prev = sum(item["sessions_prev"] or 0 for item in summaries)
    total_clicks = sum(item["event_clicks"] or 0 for item in summaries)
    avg_plp_conv = average([item["plp_conv"] for item in summaries])
    avg_purchase_conv = average([item["purchase_conv"] for item in summaries])

    ranked_plp = sorted(
        [
            [f'{value["category"]} {value["country"]}', summary["plp_conv"]]
            for key, value in parsed_sheets.items()
            for summary in [sheet_summaries[key]]
            if summary["plp_conv"] is not None
        ],
        key=lambda row: row[1],
        reverse=True,
    )
    return {
        "latest_month": latest,
        "prev_month": prev,
        "total_sessions": total_sessions,
        "session_change_pct": pct_change(total_sessions, total_sessions_prev),
        "total_clicks": total_clicks,
        "avg_plp_conv": avg_plp_conv,
        "avg_purchase_conv": avg_purchase_conv,
        "total_pages_tracked": len(parsed_sheets),
        "top_plp_pages": ranked_plp[:3],
        "low_plp_pages": ranked_plp[-3:],
    }


def build_insights(parsed_sheets, sheet_summaries, country_reports):
    items = []
    by_plp = sorted(
        [
            (key, summary["plp_conv"])
            for key, summary in sheet_summaries.items()
            if summary["plp_conv"] is not None
        ],
        key=lambda item: item[1],
        reverse=True,
    )
    if by_plp:
        best_key, best_plp = by_plp[0]
        best_sheet = parsed_sheets[best_key]
        items.append({
            "type": "success",
            "page": f'{best_sheet["category"]} {best_sheet["country"]}',
            "metric": "Conversion Efficiency",
            "content_type": best_sheet["content_type"],
            "priority": "high",
            "message": f'{best_sheet["category"]} {best_sheet["country"]} leads PLP conversion at {(best_plp * 100):.1f}%. Use this page as a benchmark for layout and CTA structure.',
        })

    for key, summary in sheet_summaries.items():
        organic = summary["organic"] or 0
        external = summary["external"] or 0
        if external > 0 and organic / external < 0.3:
            item = parsed_sheets[key]
            items.append({
                "type": "warning",
                "page": f'{item["category"]} {item["country"]}',
                "metric": "SEO",
                "content_type": item["content_type"],
                "priority": "medium",
                "message": f'{item["category"]} {item["country"]}: Organic makes up only {(organic / external) * 100:.0f}% of external traffic. SEO and discoverability need attention.',
            })

    for country, report in country_reports.items():
        if (report["session_change_pct"] or 0) <= -10:
            items.append({
                "type": "warning",
                "page": country,
                "metric": "Sessions",
                "content_type": "lineup",
                "priority": "high",
                "message": f'{country} sessions are down {(report["session_change_pct"] or 0):.1f}% MoM. Check content freshness and traffic source changes.',
            })
    return items[:8]


def build_narrative(executive, country_reports, sheet_summaries):
    top_countries = sorted(country_reports.items(), key=lambda item: item[1]["total_sessions"], reverse=True)[:5]
    low_organic = []
    high_traffic_low_plp = []
    for key, summary in sheet_summaries.items():
        organic = summary["organic"] or 0
        external = summary["external"] or 0
        if external > 0 and organic / external < 0.3:
            low_organic.append(key)
        if (summary["sessions"] or 0) >= 50 and (summary["plp_conv"] or 0) < 0.2:
            high_traffic_low_plp.append(key)

    bodies = [
        {
            "body": f"Top traffic markets are {', '.join(country for country, _ in top_countries)}. Focus CRO experiments and merchandising updates where traffic density is already proven.",
            "body_ko": f"상위 트래픽 시장은 {', '.join(country for country, _ in top_countries)} 입니다. 이미 트래픽이 큰 시장에 CRO 실험과 머천다이징 개선을 우선 적용하세요.",
        },
        {
            "body": f"Low-organic pages detected in {len(low_organic)} sheet(s). Strengthen non-brand SEO, internal linking, and search-intent alignment to reduce overreliance on non-organic acquisition.",
            "body_ko": f"오가닉 비중이 낮은 시트가 {len(low_organic)}개 확인되었습니다. 비브랜드 SEO, 내부링크, 검색 의도 정렬을 강화해 비오가닉 의존도를 낮출 필요가 있습니다.",
        },
        {
            "body": f"{len(high_traffic_low_plp)} page(s) show meaningful sessions but weak PLP conversion. Rework CTA copy, product card visibility, and comparison-entry interactions first.",
            "body_ko": f"의미 있는 세션 대비 PLP 전환이 약한 페이지가 {len(high_traffic_low_plp)}개입니다. CTA 문구, 상품 카드 가시성, 비교 진입 인터랙션부터 우선 개선하세요.",
        },
    ]
    actions = []
    for template, body in zip(NARRATIVE_TEMPLATES, bodies):
        actions.append({**template, **body})
    return {
        "content_actions": actions,
        "latest_month": executive["latest_month"],
        "prev_month": executive["prev_month"],
        "total_sessions": executive["total_sessions"],
        "total_pages": executive["total_pages_tracked"],
    }


def build_payload_from_workbooks(workbook_specs):
    parsed_sheets = {}
    periods_seen = []
    ai_data = {}

    for workbook_path, content_type in workbook_specs:
        wb_values = load_workbook(workbook_path, data_only=True)
        wb_formula = load_workbook(workbook_path, data_only=False)
        if "AI" in wb_values.sheetnames and "AI" in wb_formula.sheetnames:
            ai_data[content_type] = parse_ai_sheet(wb_values["AI"], wb_formula["AI"], content_type)
        for sheet_name in wb_values.sheetnames:
            if sheet_name in IGNORED_SHEETS:
                continue
            item = parse_sheet(wb_values[sheet_name], wb_formula[sheet_name], content_type)
            parsed_sheets[f'{content_type}_{item["category"]}_{item["country"]}'] = item
            periods_seen.extend(item.pop("_periods", []))

    periods_available = []
    period_candidates = sorted(
        {period for period in periods_seen},
        key=lambda value: (int(str(value).split("-")[0]), MONTH_NAMES.index(str(value).split("-")[1])),
    )
    for period in period_candidates:
        total_sessions_for_period = 0
        has_session_data = False
        _, month = str(period).split("-", 1)
        for item in parsed_sheets.values():
            metrics = item["metrics_2026"]
            sessions = (
                get_metric_value(metrics, period, "lineup", "session")
                or get_metric_value(metrics, period, item["category"].lower(), "lineup", "session")
                or get_metric_value(metrics, period, "session")
            )
            if sessions is not None:
                has_session_data = True
                total_sessions_for_period += sessions
        if has_session_data and total_sessions_for_period > 0:
            periods_available.append(period)
    latest = periods_available[-1] if periods_available else None
    prev = periods_available[-2] if len(periods_available) > 1 else None

    sheet_summaries = {
        key: summarize_sheet(item, latest, prev)
        for key, item in parsed_sheets.items()
    }
    country_reports = build_country_reports(parsed_sheets, sheet_summaries, latest, prev)
    category_reports = build_category_reports(parsed_sheets, sheet_summaries)
    executive = build_executive(parsed_sheets, sheet_summaries, latest, prev)
    insights = build_insights(parsed_sheets, sheet_summaries, country_reports)
    narrative = build_narrative(executive, country_reports, sheet_summaries)

    payload = {
        "meta": {
            "months_available": [str(period).split("-", 1)[1] for period in periods_available if str(period).startswith(str(latest).split("-", 1)[0])],
            "periods_available": periods_available,
            "sheets_parsed": list(parsed_sheets.keys()),
            "last_updated": None,
            "countries": sorted({item["country"] for item in parsed_sheets.values()}),
            "categories": sorted({item["category"] for item in parsed_sheets.values()}),
            "content_types": ["lineup"],
            "source_workbooks": [str(path) for path, _ in workbook_specs],
            "ai_latest_months": {key: value.get("latest_month") for key, value in ai_data.items()},
        },
        "pages": [],
        "ai_data": ai_data,
        "monthly_data": parsed_sheets,
        "insights": insights,
        "expert": {
            "executive": executive,
            "country_reports": country_reports,
            "category_reports": category_reports,
            "content_types": ["lineup"],
            "per_content_type": {},
            "sheet_summaries": sheet_summaries,
        },
        "narrative": narrative,
    }

    per_content_type = {}
    for content_type in sorted({item["content_type"] for item in parsed_sheets.values()}):
        scoped_sheets = {key: item for key, item in parsed_sheets.items() if item["content_type"] == content_type}
        scoped_summaries = {key: sheet_summaries[key] for key in scoped_sheets}
        scoped_countries = build_country_reports(scoped_sheets, scoped_summaries, latest, prev)
        scoped_categories = build_category_reports(scoped_sheets, scoped_summaries)
        scoped_executive = build_executive(scoped_sheets, scoped_summaries, latest, prev)
        per_content_type[content_type] = {
            "executive": scoped_executive,
            "country_reports": scoped_countries,
            "category_reports": scoped_categories,
        }

    payload["meta"]["content_types"] = sorted(per_content_type.keys())
    payload["expert"]["content_types"] = sorted(per_content_type.keys())
    payload["expert"]["per_content_type"] = per_content_type
    return payload


def main():
    parser = argparse.ArgumentParser(description="Generate dashboard data.json from the Excel workbook.")
    parser.add_argument("--lineup-workbook", default=str(ROOT / "data" / "monthly_ga_lineupguide.xlsx"))
    parser.add_argument("--feature-workbook", default=str(ROOT / "data" / "monthly_ga_featurelibrary.xlsx"))
    parser.add_argument("--output", default=str(ROOT / "data" / "workbook-lineup.json"))
    args = parser.parse_args()

    output_path = Path(args.output)
    workbook_specs = []
    lineup_path = Path(args.lineup_workbook)
    if lineup_path.exists():
        workbook_specs.append((lineup_path, "lineup"))
    feature_path = Path(args.feature_workbook)
    if feature_path.exists():
        workbook_specs.append((feature_path, "feature_library"))
    if not workbook_specs:
        raise SystemExit("No workbook files found.")

    payload = build_payload_from_workbooks(workbook_specs)
    output_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Wrote {output_path}")


if __name__ == "__main__":
    main()
